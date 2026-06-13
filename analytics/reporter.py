"""
reporter.py
-----------
Developer-built: Auto-generate Excel and text reports from analytics results.
Combines EDA profile, AI insights, sentiment, and NLP into downloadable reports.
No AI involved here — pure report formatting and file generation.
"""

import io
import json
import logging
from datetime import datetime
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)


def generate_excel_report(
    df: pd.DataFrame,
    analysis_result: dict,
    url: str = "",
) -> bytes:
    """
    Generate a multi-sheet Excel report with:
    - Sheet 1: Raw extracted data
    - Sheet 2: EDA profile stats
    - Sheet 3: AI insights and recommendations
    - Sheet 4: Keyword frequency
    - Sheet 5: Sentiment results

    Developer-built: Excel formatting using openpyxl via pandas.

    Args:
        df: Extracted data DataFrame.
        analysis_result: Full result dict from DataAnalyticsAgent.
        url: Source URL.

    Returns:
        Excel file as bytes (for Streamlit download button).
    """
    output = io.BytesIO()

    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # ── Sheet 1: Raw Data ─────────────────────────────────
            df.to_excel(writer, sheet_name="Raw Data", index=False)
            ws = writer.sheets["Raw Data"]
            ws.column_dimensions["A"].width = 30
            for col in ws.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.font = _bold_font()
                    cell.fill = _header_fill()

            # ── Sheet 2: EDA Profile ──────────────────────────────
            profile = analysis_result.get("profile", {})
            profile_rows = []

            # Shape
            shape = profile.get("shape", {})
            profile_rows.append({"Section": "Shape", "Metric": "Total Rows", "Value": shape.get("rows", 0)})
            profile_rows.append({"Section": "Shape", "Metric": "Total Columns", "Value": shape.get("columns", 0)})
            profile_rows.append({"Section": "Shape", "Metric": "Duplicate Rows", "Value": profile.get("duplicates", 0)})
            profile_rows.append({"Section": "Quality", "Metric": "Quality Score", "Value": f"{analysis_result.get('data_quality_score', 0)}/100"})

            # Missing values
            for col, miss in profile.get("missing", {}).items():
                if miss["count"] > 0:
                    profile_rows.append({
                        "Section": "Missing Values",
                        "Metric": col,
                        "Value": f"{miss['count']} ({miss['percent']}%)",
                    })

            # Numeric stats
            for col, stats in profile.get("numeric_stats", {}).items():
                for metric, val in stats.items():
                    profile_rows.append({
                        "Section": f"Numeric — {col}",
                        "Metric": metric,
                        "Value": val,
                    })

            # Categorical stats
            for col, stats in profile.get("categorical_stats", {}).items():
                profile_rows.append({
                    "Section": f"Categorical — {col}",
                    "Metric": "Unique Values",
                    "Value": stats.get("unique_count", 0),
                })
                profile_rows.append({
                    "Section": f"Categorical — {col}",
                    "Metric": "Most Frequent",
                    "Value": stats.get("most_frequent", ""),
                })

            if profile_rows:
                profile_df = pd.DataFrame(profile_rows)
                profile_df.to_excel(writer, sheet_name="EDA Profile", index=False)
                ws2 = writer.sheets["EDA Profile"]
                for col in ws2.iter_cols(min_row=1, max_row=1):
                    for cell in col:
                        cell.font = _bold_font()
                        cell.fill = _header_fill()

            # ── Sheet 3: AI Insights ──────────────────────────────
            insights_data = {
                "Section": ["Source URL", "Analysis Date", "AI Insights", "Recommendations"],
                "Content": [
                    url,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    analysis_result.get("ai_insights", ""),
                    analysis_result.get("recommendations", ""),
                ],
            }
            insights_df = pd.DataFrame(insights_data)
            insights_df.to_excel(writer, sheet_name="AI Insights", index=False)
            ws3 = writer.sheets["AI Insights"]
            ws3.column_dimensions["A"].width = 20
            ws3.column_dimensions["B"].width = 80
            for col in ws3.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.font = _bold_font()
                    cell.fill = _header_fill()

            # ── Sheet 4: Keywords ─────────────────────────────────
            keywords = analysis_result.get("keywords", [])
            if keywords:
                kw_df = pd.DataFrame(keywords)
                kw_df.to_excel(writer, sheet_name="Keywords", index=False)
                ws4 = writer.sheets["Keywords"]
                for col in ws4.iter_cols(min_row=1, max_row=1):
                    for cell in col:
                        cell.font = _bold_font()
                        cell.fill = _header_fill()

            # ── Sheet 5: Sentiment ────────────────────────────────
            sentiment = analysis_result.get("sentiment", {})
            if sentiment:
                overall = sentiment.get("overall", {})
                sent_rows = [
                    {"Metric": "Overall Label", "Value": overall.get("label", "")},
                    {"Metric": "Polarity Score", "Value": overall.get("polarity", 0)},
                    {"Metric": "Subjectivity", "Value": overall.get("subjectivity", 0)},
                    {"Metric": "Emoji", "Value": overall.get("emoji", "")},
                    {"Metric": "Sentence Count", "Value": sentiment.get("sentence_count", 0)},
                    {"Metric": "Avg Polarity", "Value": sentiment.get("average_polarity", 0)},
                ]
                sent_df = pd.DataFrame(sent_rows)
                sent_df.to_excel(writer, sheet_name="Sentiment", index=False)
                ws5 = writer.sheets["Sentiment"]
                for col in ws5.iter_cols(min_row=1, max_row=1):
                    for cell in col:
                        cell.font = _bold_font()
                        cell.fill = _header_fill()

        output.seek(0)
        logger.info("[Reporter] Excel report generated successfully")
        return output.getvalue()

    except Exception as e:
        logger.error(f"[Reporter] Excel generation failed: {e}")
        return b""


def generate_text_report(
    analysis_result: dict,
    url: str = "",
    content_type: str = "",
) -> str:
    """
    Generate a plain-text analytics report.

    Args:
        analysis_result: Full result from DataAnalyticsAgent.
        url: Source URL.
        content_type: Type of scraped content.

    Returns:
        Formatted text report string.
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    profile = analysis_result.get("profile", {})
    shape = profile.get("shape", {})

    lines = [
        "=" * 70,
        "  AUTONOMOUS DATA ANALYTICS REPORT",
        "  Web Scraping AI Bot — Powered by Gemini",
        "=" * 70,
        f"  Generated : {now}",
        f"  Source URL: {url}",
        f"  Data Type : {content_type or 'Auto-detected'}",
        "=" * 70,
        "",
        "── DATASET OVERVIEW ──────────────────────────────────────",
        f"  Rows      : {shape.get('rows', 0)}",
        f"  Columns   : {shape.get('columns', 0)}",
        f"  Duplicates: {profile.get('duplicates', 0)}",
        f"  Quality   : {analysis_result.get('data_quality_score', 0)}/100",
        "",
    ]

    # Missing values
    missing = {k: v for k, v in profile.get("missing", {}).items() if v["count"] > 0}
    if missing:
        lines.append("── MISSING VALUES ────────────────────────────────────────")
        for col, m in missing.items():
            lines.append(f"  {col}: {m['count']} missing ({m['percent']}%)")
        lines.append("")

    # Numeric stats
    numeric_stats = profile.get("numeric_stats", {})
    if numeric_stats:
        lines.append("── NUMERIC STATISTICS ────────────────────────────────────")
        for col, stats in numeric_stats.items():
            lines.append(f"  {col}:")
            lines.append(f"    Mean: {stats.get('mean')}  |  Median: {stats.get('median')}")
            lines.append(f"    Min : {stats.get('min')}  |  Max   : {stats.get('max')}")
            lines.append(f"    Std : {stats.get('std')}  |  Outliers (IQR): {stats.get('outliers_iqr', 0)}")
        lines.append("")

    # Sentiment
    sentiment = analysis_result.get("sentiment", {})
    if sentiment and sentiment.get("overall"):
        overall = sentiment["overall"]
        lines.append("── SENTIMENT ANALYSIS ────────────────────────────────────")
        lines.append(f"  Overall  : {overall.get('emoji')} {overall.get('label', '').upper()}")
        lines.append(f"  Polarity : {overall.get('polarity')} (range: -1 to +1)")
        lines.append(f"  Subjectivity: {overall.get('subjectivity')} (0=objective, 1=subjective)")
        lines.append("")

    # Top keywords
    keywords = analysis_result.get("keywords", [])
    if keywords:
        lines.append("── TOP KEYWORDS ──────────────────────────────────────────")
        kw_list = [f"{k['word']} ({k['count']})" for k in keywords[:15]]
        lines.append("  " + " | ".join(kw_list))
        lines.append("")

    # Topics
    topics = analysis_result.get("topics", [])
    if topics:
        lines.append("── DETECTED TOPICS ───────────────────────────────────────")
        for t in topics[:5]:
            bar = "█" * int(t["confidence"] * 20)
            lines.append(f"  {t['topic']:<20} {bar} {t['confidence']:.1%}")
        lines.append("")

    # AI insights
    ai_insights = analysis_result.get("ai_insights", "")
    if ai_insights:
        lines.append("── AI INSIGHTS (Gemini) ──────────────────────────────────")
        lines.append("")
        lines.append(ai_insights)
        lines.append("")

    # Recommendations
    recommendations = analysis_result.get("recommendations", "")
    if recommendations:
        lines.append("── AI RECOMMENDATIONS (Gemini) ───────────────────────────")
        lines.append("")
        lines.append(recommendations)
        lines.append("")

    lines.append("=" * 70)
    lines.append("  END OF REPORT — Web Scraping AI Bot")
    lines.append("=" * 70)

    return "\n".join(lines)


def _bold_font():
    """Return openpyxl bold font."""
    try:
        from openpyxl.styles import Font
        return Font(bold=True, color="FFFFFF")
    except ImportError:
        return None


def _header_fill():
    """Return openpyxl header fill color."""
    try:
        from openpyxl.styles import PatternFill
        return PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    except ImportError:
        return None
